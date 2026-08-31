"""Builds the Excel deliverable.

Four sheets: an executive Dashboard (KPI row + three charts), the full Data
sheet, a per-category Summary with in-cell data bars, and a Quality sheet.
"""

import statistics
from collections import Counter, defaultdict
from typing import Dict, List

from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

DARK = "1F4E79"
ACCENT = "2E86AB"
LIGHT = "DDEBF7"
HEADER_FILL = PatternFill("solid", fgColor=DARK)
HEADER_FONT = Font(color="FFFFFF", bold=True)
KPI_FILL = PatternFill("solid", fgColor=LIGHT)
THIN_BORDER = Border(*[Side(style="thin", color="B7C9E0")] * 4)


def _style_header(sheet, columns_count: int, row: int = 1) -> None:
    for col in range(1, columns_count + 1):
        cell = sheet.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    sheet.freeze_panes = sheet.cell(row=row + 1, column=1).coordinate


def _fit_columns(sheet, widths: List[int]) -> None:
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width


def _kpi_box(sheet, column: int, label: str, value, number_format=None) -> None:
    """Two stacked merged cells styled as a KPI tile (3 columns wide)."""
    left, right = column, column + 2
    for row, content in ((3, label), (4, value)):
        sheet.merge_cells(start_row=row, start_column=left, end_row=row, end_column=right)
        cell = sheet.cell(row=row, column=left)
        cell.value = content
        cell.alignment = Alignment(horizontal="center", vertical="center")
        for c in range(left, right + 1):
            sheet.cell(row=row, column=c).fill = KPI_FILL
            sheet.cell(row=row, column=c).border = THIN_BORDER
    sheet.cell(row=3, column=left).font = Font(size=10, color="44546A")
    value_cell = sheet.cell(row=4, column=left)
    value_cell.font = Font(size=16, bold=True, color=DARK)
    if number_format:
        value_cell.number_format = number_format


def _bar(title: str, data_ref: Reference, cats_ref: Reference, color: str) -> BarChart:
    chart = BarChart()
    chart.type = "col"
    chart.title = title
    chart.height, chart.width = 8, 15
    chart.legend = None
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.series[0].graphicalProperties.solidFill = color
    return chart


def build_report(rows: List[Dict], quality: Dict[str, int], out_path: str) -> None:
    workbook = Workbook()
    prices = [row["price"] for row in rows]
    by_category = defaultdict(list)
    for row in rows:
        by_category[row["category"]].append(row)

    # ---- Dashboard ------------------------------------------------------
    dash = workbook.active
    dash.title = "Dashboard"
    dash.sheet_view.showGridLines = False
    dash.merge_cells("A1:R1")
    dash["A1"] = "Catalog analytics — books.toscrape.com"
    dash["A1"].font = Font(size=18, bold=True, color=DARK)

    in_stock_share = sum(1 for r in rows if r["in_stock"]) / len(rows)
    top_rated_share = sum(1 for r in rows if r["rating"] >= 4) / len(rows)
    kpis = [
        ("BOOKS", len(rows), None),
        ("CATEGORIES", len(by_category), None),
        ("AVG PRICE", round(sum(prices) / len(prices), 2), '£#,##0.00'),
        ("MEDIAN PRICE", round(statistics.median(prices), 2), '£#,##0.00'),
        ("IN STOCK", in_stock_share, "0%"),
        ("RATED 4-5", top_rated_share, "0%"),
    ]
    for index, (label, value, fmt) in enumerate(kpis):
        _kpi_box(dash, 1 + index * 3, label, value, fmt)

    # Helper series live on a separate Calc sheet so the dashboard stays clean.
    calc = workbook.create_sheet("Calc")
    top_categories = sorted(by_category.items(), key=lambda kv: -len(kv[1]))[:10]
    calc.append(["Category", "Books"])
    for name, items in top_categories:
        calc.append([name, len(items)])

    bins = [(0, 20), (20, 30), (30, 40), (40, 50), (50, 100)]
    calc.append([]); hist_start = calc.max_row + 1
    calc.append(["Price band", "Books"])
    for low, high in bins:
        label = "£%d-%d" % (low, high)
        calc.append([label, sum(1 for p in prices if low <= p < high)])

    rating_counts = Counter(r["rating"] for r in rows)
    calc.append([]); rating_start = calc.max_row + 1
    calc.append(["Rating", "Books"])
    for stars in sorted(rating_counts):
        calc.append(["%d star%s" % (stars, "s" if stars > 1 else ""), rating_counts[stars]])
    calc.sheet_state = "hidden"

    dash.add_chart(
        _bar("Top 10 categories by titles",
             Reference(calc, min_col=2, min_row=1, max_row=1 + len(top_categories)),
             Reference(calc, min_col=1, min_row=2, max_row=1 + len(top_categories)),
             ACCENT),
        "A7",
    )
    dash.add_chart(
        _bar("Price distribution",
             Reference(calc, min_col=2, min_row=hist_start, max_row=hist_start + len(bins)),
             Reference(calc, min_col=1, min_row=hist_start + 1, max_row=hist_start + len(bins)),
             "F18F01"),
        "H7",
    )
    donut = DoughnutChart()
    donut.title = "Rating distribution"
    donut.height, donut.width = 8, 11
    donut.add_data(
        Reference(calc, min_col=2, min_row=rating_start, max_row=rating_start + len(rating_counts)),
        titles_from_data=True,
    )
    donut.set_categories(
        Reference(calc, min_col=1, min_row=rating_start + 1, max_row=rating_start + len(rating_counts))
    )
    donut.dataLabels = DataLabelList()
    donut.dataLabels.showPercent = True
    dash.add_chart(donut, "O7")

    # ---- Data -----------------------------------------------------------
    data = workbook.create_sheet("Data")
    data.append(["Title", "Category", "Price", "Rating", "In stock", "URL"])
    for row in sorted(rows, key=lambda r: (r["category"], r["title"])):
        data.append([row["title"], row["category"], row["price"], row["rating"],
                     "yes" if row["in_stock"] else "no", row["url"]])
    _style_header(data, 6)
    _fit_columns(data, [52, 24, 10, 8, 9, 60])
    price_range = "C2:C%d" % data.max_row
    for cell_row in data[price_range]:
        cell_row[0].number_format = "£0.00"
    data.conditional_formatting.add(
        price_range,
        ColorScaleRule(start_type="min", start_color="C6E0B4",
                       end_type="max", end_color="F8696B"),
    )

    # ---- Summary --------------------------------------------------------
    summary = workbook.create_sheet("Summary")
    summary.append(["Category", "Books", "Avg price", "Min price", "Max price", "4-5 star share"])
    for category in sorted(by_category, key=lambda c: -len(by_category[c])):
        items = by_category[category]
        cat_prices = [item["price"] for item in items]
        summary.append([
            category, len(items), sum(cat_prices) / len(cat_prices),
            min(cat_prices), max(cat_prices),
            sum(1 for item in items if item["rating"] >= 4) / len(items),
        ])
    _style_header(summary, 6)
    _fit_columns(summary, [24, 8, 10, 10, 10, 14])
    last = summary.max_row
    for cell_row in summary["C2:E%d" % last]:
        for cell in cell_row:
            cell.number_format = "£0.00"
    for cell_row in summary["F2:F%d" % last]:
        cell_row[0].number_format = "0%"
    summary.conditional_formatting.add(
        "B2:B%d" % last,
        DataBarRule(start_type="num", start_value=0, end_type="max",
                    color=ACCENT, showValue=True),
    )
    summary.conditional_formatting.add(
        "F2:F%d" % last,
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1,
                    color="F18F01", showValue=True),
    )

    # ---- Quality --------------------------------------------------------
    notes = workbook.create_sheet("Quality")
    notes.append(["Metric", "Value"])
    for key in ("total_raw", "clean", "duplicates", "bad_price", "bad_rating"):
        notes.append([key.replace("_", " "), quality.get(key, 0)])
    _style_header(notes, 2)
    _fit_columns(notes, [18, 10])

    workbook.save(out_path)
